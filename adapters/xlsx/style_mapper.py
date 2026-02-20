"""Maps core styles to openpyxl styles."""

from __future__ import annotations

from typing import Optional

from openpyxl.styles import Alignment as OpenpyxlAlignment
from openpyxl.styles import Border as OpenpyxlBorder
from openpyxl.styles import Font as OpenpyxlFont
from openpyxl.styles import PatternFill as OpenpyxlPatternFill
from openpyxl.styles import Side as OpenpyxlSide

from ...core.styles.cell_style import (
    Alignment,
    Border,
    BorderStyle,
    CellStyle,
    Fill,
    Font,
    HorizontalAlignment,
    PatternFill,
    UnderlineStyle,
    VerticalAlignment,
)


class StyleMapper:
    """Maps core style objects to openpyxl style objects."""

    @staticmethod
    def map_cell_style(style: Optional[CellStyle]) -> dict:
        """
        Map CellStyle to openpyxl style components.

        Returns:
            Dictionary with font, fill, border, alignment, number_format
        """
        if style is None:
            return {}

        result = {}

        if style.font:
            result['font'] = StyleMapper.map_font(style.font)

        if style.fill:
            result['fill'] = StyleMapper.map_fill(style.fill)

        if style.border:
            result['border'] = StyleMapper.map_border(style.border)

        if style.alignment:
            result['alignment'] = StyleMapper.map_alignment(style.alignment)

        if style.number_format:
            result['number_format'] = style.number_format

        return result

    @staticmethod
    def map_font(font: Font) -> OpenpyxlFont:
        """Map Font to openpyxl Font."""
        kwargs = {}

        if font.family:
            kwargs['name'] = font.family

        if font.size:
            kwargs['size'] = font.size

        kwargs['bold'] = font.bold
        kwargs['italic'] = font.italic

        if font.underline != UnderlineStyle.NONE:
            kwargs['underline'] = font.underline.value

        kwargs['strike'] = font.strikethrough

        if font.color:
            kwargs['color'] = str(font.color)

        return OpenpyxlFont(**kwargs)

    @staticmethod
    def map_fill(fill: Fill) -> OpenpyxlPatternFill:
        """Map Fill to openpyxl PatternFill."""
        kwargs = {}

        if fill.pattern != PatternFill.NONE:
            kwargs['patternType'] = fill.pattern.value

        if fill.foreground_color:
            kwargs['fgColor'] = str(fill.foreground_color)

        if fill.background_color:
            kwargs['bgColor'] = str(fill.background_color)

        return OpenpyxlPatternFill(**kwargs)

    @staticmethod
    def map_border(border: Border) -> OpenpyxlBorder:
        """Map Border to openpyxl Border."""
        kwargs = {}

        if border.left:
            kwargs['left'] = StyleMapper._map_border_side(border.left, border.color)

        if border.right:
            kwargs['right'] = StyleMapper._map_border_side(border.right, border.color)

        if border.top:
            kwargs['top'] = StyleMapper._map_border_side(border.top, border.color)

        if border.bottom:
            kwargs['bottom'] = StyleMapper._map_border_side(border.bottom, border.color)

        return OpenpyxlBorder(**kwargs)

    @staticmethod
    def _map_border_side(style: BorderStyle, color) -> OpenpyxlSide:
        """Map BorderStyle to openpyxl Side."""
        kwargs = {}

        if style != BorderStyle.NONE:
            kwargs['style'] = style.value

        if color:
            kwargs['color'] = str(color)

        return OpenpyxlSide(**kwargs)

    @staticmethod
    def map_alignment(alignment: Alignment) -> OpenpyxlAlignment:
        """Map Alignment to openpyxl Alignment."""
        kwargs = {}

        if alignment.horizontal:
            kwargs['horizontal'] = alignment.horizontal.value

        if alignment.vertical:
            kwargs['vertical'] = alignment.vertical.value

        kwargs['wrap_text'] = alignment.wrap_text
        kwargs['shrink_to_fit'] = alignment.shrink_to_fit

        if alignment.indent > 0:
            kwargs['indent'] = alignment.indent

        if alignment.text_rotation != 0:
            kwargs['text_rotation'] = alignment.text_rotation

        return OpenpyxlAlignment(**kwargs)
