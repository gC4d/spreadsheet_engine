"""XLSX adapter using openpyxl."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ...core.models.cell import Cell, CellDataType
from ...core.models.position import CellPosition
from ...core.models.sheet import Sheet
from ...core.models.spreadsheet import Spreadsheet
from ...core.models.table import Table
from .style_mapper import StyleMapper


class XLSXAdapter:
    """
    XLSX adapter using openpyxl.

    Translates core spreadsheet models to openpyxl workbooks.
    """

    def render(self, spreadsheet: Spreadsheet, autofit: bool = True) -> Workbook:
        """Render spreadsheet to openpyxl Workbook."""
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        for sheet_model in spreadsheet.sheets:
            ws = workbook.create_sheet(title=sheet_model.name)
            self._render_sheet(ws, sheet_model, autofit)

        if spreadsheet.active_sheet:
            try:
                workbook.active = workbook[spreadsheet.active_sheet]
            except KeyError:
                pass

        return workbook

    def _render_sheet(self, ws: Any, sheet: Sheet, autofit: bool) -> None:
        """Render a sheet to openpyxl worksheet."""
        for table in sheet.tables:
            self._render_table(ws, table)

        for position, cell in sheet.cells.items():
            self._write_cell(ws, cell, position)

        if sheet.freeze_panes:
            freeze_cell = sheet.freeze_panes.to_a1()
            ws.freeze_panes = freeze_cell

        for column, width in sheet.column_widths.items():
            col_letter = get_column_letter(column)
            ws.column_dimensions[col_letter].width = width

        for row, height in sheet.row_heights.items():
            ws.row_dimensions[row].height = height

        if autofit:
            self._autofit_columns(ws)

    def _render_table(self, ws: Any, table: Table) -> None:
        """Render a table to worksheet."""
        current_row = table.start_position.row
        start_col = table.start_position.column

        if table.has_title and table.title:
            self._write_cell(
                ws,
                table.title,
                CellPosition(current_row, start_col),
            )

            if table.column_count > 1:
                end_col = start_col + table.column_count - 1
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                ws.merge_cells(f"{start_letter}{current_row}:{end_letter}{current_row}")

            current_row += 1

        if table.has_headers:
            for col_idx, header_cell in enumerate(table.headers):
                col_position = start_col + col_idx
                self._write_cell(
                    ws,
                    header_cell,
                    CellPosition(current_row, col_position),
                )
            current_row += 1

        for row_cells in table.iter_rows():
            for col_idx, cell in enumerate(row_cells):
                col_position = start_col + col_idx
                self._write_cell(
                    ws,
                    cell,
                    CellPosition(current_row, col_position),
                )
            current_row += 1

    def _write_cell(self, ws: Any, cell: Cell, position: CellPosition) -> None:
        """Write a cell to worksheet."""
        openpyxl_cell = ws.cell(row=position.row, column=position.column)

        if cell.is_formula and cell.formula:
            openpyxl_cell.value = cell.formula
        else:
            openpyxl_cell.value = cell.value

        if cell.number_format:
            openpyxl_cell.number_format = cell.number_format

        if cell.style:
            style_components = StyleMapper.map_cell_style(cell.style)

            if 'font' in style_components:
                openpyxl_cell.font = style_components['font']

            if 'fill' in style_components:
                openpyxl_cell.fill = style_components['fill']

            if 'border' in style_components:
                openpyxl_cell.border = style_components['border']

            if 'alignment' in style_components:
                openpyxl_cell.alignment = style_components['alignment']

        if cell.hyperlink:
            openpyxl_cell.hyperlink = cell.hyperlink

        if cell.comment:
            from openpyxl.comments import Comment
            openpyxl_cell.comment = Comment(cell.comment, "System")

    def _autofit_columns(self, ws: Any) -> None:
        """Auto-fit column widths based on content."""
        column_widths = {}

        for row in ws.iter_rows():
            for cell in row:
                column_letter = getattr(cell, "column_letter", None)
                if column_letter is None:
                    continue

                if cell.value is None:
                    continue

                cell_length = len(str(cell.value))
                column_widths[column_letter] = max(
                    column_widths.get(column_letter, 0), cell_length
                )

        for column_letter, max_length in column_widths.items():
            adjusted_width = max_length + 2
            adjusted_width = max(10, min(50, adjusted_width))
            ws.column_dimensions[column_letter].width = adjusted_width

    def to_bytes(self, workbook: Workbook) -> bytes:
        """Convert workbook to bytes."""
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def to_file(self, workbook: Workbook, path: Path) -> None:
        """Save workbook to file."""
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(path))

    def to_stream(self, workbook: Workbook, stream: BytesIO) -> None:
        """Write workbook to stream."""
        workbook.save(stream)
        stream.seek(0)
